# Copyright (c) 2026 Waterfall Project
# SPDX-License-Identifier: Commercial

"""Excel parsers for expenses and RAE data."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from poc_import.models import (
    ExcelExpensesData,
    ExcelRAEData,
    ExpenseEntry,
    ExpenseRow,
    RAEEntry,
    RAETaskBreakdown,
)

EXPENSES_SHEET_NAME = "Dépenses"
RAE_SHEET_NAME = "RAE Forecast"

REQUIRED_EXPENSE_COLUMNS = {
    "Exercice comptable",
    "Période",
    "Elément d'OTP",
    "Désign.nat.comptable",
    "Val./Devise objet",
    "Date de la pièce",
}
REQUIRED_RAE_COLUMNS = {"milestone_name", "remaining_amount", "forecast_date"}


def parse_expenses_excel(path: Path) -> ExcelExpensesData:
    """Parse expenses Excel file into structured rows and grouped entries.

    Args:
        path: Path to Excel file.

    Returns:
        Parsed expenses data with grouping summary.

    Raises:
        ValueError: If required columns are missing.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = _get_sheet(workbook, EXPENSES_SHEET_NAME)
    header_map = _get_header_map(sheet)
    missing = REQUIRED_EXPENSE_COLUMNS - set(header_map.keys())
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_list}")

    rows: list[ExpenseRow] = []
    for row_index, row_values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if _row_is_empty(row_values):
            continue
        row = ExpenseRow(
            row_number=row_index,
            purchase_document=_as_str(
                _get_cell(row_values, header_map, "Document d'achat")
            ),
            fiscal_year=_parse_int(
                _get_cell(row_values, header_map, "Exercice comptable")
            ),
            period=_parse_int(_get_cell(row_values, header_map, "Période")),
            otp_element=_as_str(_get_cell(row_values, header_map, "Elément d'OTP")),
            resource_name=_as_str(_get_cell(row_values, header_map, "Nom Matricule")),
            vendor_name=_as_str(_get_cell(row_values, header_map, "Nom 1")),
            accounting_nature=_as_str(
                _get_cell(row_values, header_map, "Nature comptable")
            ),
            accounting_nature_label=_as_str(
                _get_cell(row_values, header_map, "Désign.nat.comptable")
            ),
            reference_number=_as_str(
                _get_cell(row_values, header_map, "Nº pièce référence")
            ),
            amount=_parse_float(_get_cell(row_values, header_map, "Val./Devise objet")),
            expense_date=_parse_date(
                _get_cell(row_values, header_map, "Date de la pièce")
            ),
            description=_as_str(
                _get_cell(row_values, header_map, "Texte de la commande d'achat")
            ),
            origin_group=_as_str(_get_cell(row_values, header_map, "Groupe d'origine")),
            purchase_reference=_as_str(_get_cell(row_values, header_map, "Référence")),
        )
        rows.append(row)

    entries, summary = _group_expense_rows(rows)
    return ExcelExpensesData(
        file_path=str(path),
        sheet_name=sheet.title,
        rows=rows,
        entries=entries,
        total_rows=summary["total_rows"],
        total_amount=summary["total_amount"],
        period_start=summary["period_start"],
        period_end=summary["period_end"],
        unique_references=summary["unique_references"],
        grouped_count=summary["grouped_count"],
        missing_references=summary["missing_references"],
        purchase_rows=summary["purchase_rows"],
        purchase_amount=summary["purchase_amount"],
        time_rows=summary["time_rows"],
        time_amount=summary["time_amount"],
    )


def parse_rae_excel(path: Path) -> ExcelRAEData:
    """Parse RAE Excel file into structured entries.

    Args:
        path: Path to Excel file.

    Returns:
        Parsed RAE data.

    Raises:
        ValueError: If required columns are missing.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = _get_sheet(workbook, RAE_SHEET_NAME)
    header_map = _get_header_map(sheet)
    missing = REQUIRED_RAE_COLUMNS - set(header_map.keys())
    if missing:
        missing_list = ", ".join(sorted(missing))
        raise ValueError(f"Missing required columns: {missing_list}")

    entries: list[RAEEntry] = []
    entry_id = 1
    for row_index, row_values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if _row_is_empty(row_values):
            continue

        milestone_name = _as_str(_get_cell(row_values, header_map, "milestone_name"))
        remaining_amount = _parse_float(
            _get_cell(row_values, header_map, "remaining_amount")
        )
        forecast_date = _parse_date(_get_cell(row_values, header_map, "forecast_date"))
        breakdown_value = _get_cell(row_values, header_map, "task_breakdown")
        breakdown_items, breakdown_sum, parse_error = _parse_task_breakdown(
            breakdown_value
        )

        entries.append(
            RAEEntry(
                entry_id=entry_id,
                milestone_name=milestone_name or "",
                remaining_amount=remaining_amount,
                forecast_date=forecast_date,
                task_breakdown=breakdown_items,
                breakdown_sum=breakdown_sum,
                row_number=row_index,
                parse_error=parse_error,
            )
        )
        entry_id += 1

    forecast_period = _format_forecast_period(entries)
    total_remaining = sum(entry.remaining_amount or 0.0 for entry in entries)
    milestone_count = len(
        {entry.milestone_name for entry in entries if entry.milestone_name}
    )

    return ExcelRAEData(
        file_path=str(path),
        sheet_name=sheet.title,
        entries=entries,
        total_rows=len(entries),
        total_remaining=total_remaining,
        milestone_count=milestone_count,
        forecast_period=forecast_period,
    )


def _get_sheet(workbook: Any, preferred_name: str) -> Any:
    """Get worksheet by name or return first sheet."""
    if preferred_name in workbook.sheetnames:
        return workbook[preferred_name]
    return workbook[workbook.sheetnames[0]]


def _get_header_map(sheet: Any) -> dict[str, int]:
    """Extract column header mappings from first row."""
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        if value is None:
            continue
        header = str(value).strip()
        if header:
            header_map[header] = index
    return header_map


def _row_is_empty(row_values: tuple[Any, ...]) -> bool:
    return all(
        value is None or (isinstance(value, str) and not value.strip())
        for value in row_values
    )


def _get_cell(
    row_values: tuple[Any, ...], header_map: dict[str, int], column: str
) -> Any:
    index = header_map.get(column)
    if index is None:
        return None
    return row_values[index]


def _as_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return int(text)
        except ValueError:
            return None
    return None


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        normalized = text.replace(" ", "").replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            excel_date: date = from_excel(value).date()
            return excel_date
        except (TypeError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            return None
    return None


def _group_expense_rows(
    rows: list[ExpenseRow],
) -> tuple[list[ExpenseEntry], dict[str, Any]]:
    grouped: dict[str, list[ExpenseRow]] = {}
    for row in rows:
        reference = row.reference_number.strip() if row.reference_number else ""
        if reference:
            grouped.setdefault(reference, []).append(row)
        else:
            grouped[f"__row_{row.row_number}"] = [row]

    entries: list[ExpenseEntry] = []
    entry_id = 1
    for key, group_rows in grouped.items():
        first = group_rows[0]
        amount = sum((row.amount or 0.0) for row in group_rows)
        category = _infer_expense_category(first)
        description = _infer_expense_description(first)
        entries.append(
            ExpenseEntry(
                entry_id=entry_id,
                reference_number=(
                    first.reference_number if not key.startswith("__row_") else None
                ),
                expense_date=first.expense_date,
                amount=amount,
                category=category,
                description=description,
                purchase_document=first.purchase_document,
                fiscal_year=first.fiscal_year,
                period=first.period,
                otp_element=first.otp_element,
                resource_name=first.resource_name,
                vendor_name=first.vendor_name,
                accounting_nature=first.accounting_nature,
                accounting_nature_label=first.accounting_nature_label,
                origin_group=first.origin_group,
                purchase_reference=first.purchase_reference,
                row_numbers=[row.row_number for row in group_rows],
                grouped_rows=len(group_rows),
            )
        )
        entry_id += 1

    total_amount = sum((row.amount or 0.0) for row in rows)
    dates = [row.expense_date for row in rows if row.expense_date]
    period_start = min(dates) if dates else None
    period_end = max(dates) if dates else None
    references = [row.reference_number.strip() for row in rows if row.reference_number]
    unique_references = len(set(references))
    grouped_count = sum(1 for ref in set(references) if references.count(ref) > 1)
    missing_references = sum(1 for row in rows if not row.reference_number)

    purchase_rows = [row for row in rows if row.purchase_document]
    time_rows = [row for row in rows if not row.purchase_document]
    purchase_amount = sum((row.amount or 0.0) for row in purchase_rows)
    time_amount = sum((row.amount or 0.0) for row in time_rows)

    summary = {
        "total_rows": len(rows),
        "total_amount": total_amount,
        "period_start": period_start,
        "period_end": period_end,
        "unique_references": unique_references,
        "grouped_count": grouped_count,
        "missing_references": missing_references,
        "purchase_rows": len(purchase_rows),
        "purchase_amount": purchase_amount,
        "time_rows": len(time_rows),
        "time_amount": time_amount,
    }
    return entries, summary


def _infer_expense_category(row: ExpenseRow) -> str:
    label = (row.accounting_nature_label or "").lower()
    if not row.purchase_document:
        return "labor"
    if "sous-trait" in label or "subcontract" in label:
        return "subcontracting"
    if "frais" in label or "overhead" in label:
        return "overhead"
    if "main d'oeuvre" in label or "main d’uvre" in label:
        return "labor"
    return "procurement"


def _infer_expense_description(row: ExpenseRow) -> str | None:
    if row.description:
        return row.description
    if row.resource_name:
        return row.resource_name
    if row.vendor_name:
        return row.vendor_name
    return None


def _parse_task_breakdown(
    value: Any,
) -> tuple[list[RAETaskBreakdown], float, str | None]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return [], 0.0, None

    parsed: Any = value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError as exc:
            return [], 0.0, f"Invalid JSON: {exc}"

    if isinstance(parsed, dict) and "task_estimates" in parsed:
        parsed = parsed.get("task_estimates")

    if not isinstance(parsed, list):
        return [], 0.0, "Task breakdown must be a JSON array"

    breakdown: list[RAETaskBreakdown] = []
    breakdown_sum = 0.0
    for item in parsed:
        if not isinstance(item, dict):
            continue
        task_name = item.get("task_name") or item.get("name")
        amount = item.get("amount", item.get("remaining_cost"))
        if task_name is None or amount is None:
            continue
        amount_value = _parse_float(amount)
        if amount_value is None:
            continue
        breakdown.append(
            RAETaskBreakdown(
                task_name=str(task_name),
                amount=amount_value,
                comment=item.get("comment"),
            )
        )
        breakdown_sum += amount_value

    return breakdown, breakdown_sum, None


def _format_forecast_period(entries: list[RAEEntry]) -> str | None:
    dates = [entry.forecast_date for entry in entries if entry.forecast_date]
    if not dates:
        return None
    min_date = min(dates)
    max_date = max(dates)
    if min_date.year == max_date.year and _quarter(min_date) == _quarter(max_date):
        return f"Q{_quarter(min_date)} {min_date.year}"
    return f"{min_date.strftime('%Y-%m')} to {max_date.strftime('%Y-%m')}"


def _quarter(value: date) -> int:
    return (value.month - 1) // 3 + 1
